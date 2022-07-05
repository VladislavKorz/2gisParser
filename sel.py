from openpyxl import Workbook
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import random
import os
import requests
from loguru import logger
import base64

city_list = ['tomsk']
icon_dict = {
    'M4 12a7.83 7.83 0 0 0 8 8 8.67 8.67 0 0 0 3.41-.71l-.82-1.83A6.6 6.6 0 0 1 12 18a5.87 5.87 0 0 1-6-6 5.82 5.82 0 0 1 6.05-6A5.85 5.85 0 0 1 18 12v.5a1.5 1.5 0 0 1-3 0V8h-1.5l-.5.35A3.45 3.45 0 0 0 11.5 8 3.5 3.5 0 0 0 8 11.5v1a3.49 3.49 0 0 0 6 2.44 3.49 3.49 0 0 0 6-2.44V12a7.8 7.8 0 0 0-7.95-8A7.85 7.85 0 0 0 4 12zm7.5 2a1.5 1.5 0 0 1-1.5-1.5v-1a1.5 1.5 0 0 1 3 0v1a1.5 1.5 0 0 1-1.5 1.5': 'email',
    'M14 14l-1.08 1.45a13.61 13.61 0 0 1-4.37-4.37L10 10a18.47 18.47 0 0 0-.95-5.85L9 4H5.06a1 1 0 0 0-1 1.09 16 16 0 0 0 14.85 14.85 1 1 0 0 0 1.09-1V15h-.15A18.47 18.47 0 0 0 14 14z': 'phone',
    'M12 4a8 8 0 1 0 8 8 8 8 0 0 0-8-8zm5 9h-6l1-7h1v5.25l4 .75z': 'time_work',
    'M5 11v2a6.82 6.82 0 0 1 4.17 1.41C10.75 15.62 11.53 18 11.5 22h1c0-4 .75-6.38 2.33-7.59A6.82 6.82 0 0 1 19 13v-2a7 7 0 0 0-7-7 7 7 0 0 0-7 7z': 'address',
    'M12 4a8 8 0 1 0 8 8 8 8 0 0 0-8-8zm-6 8a5.84 5.84 0 0 1 .22-1.57L7 12h2l1 2h1v3.91A6 6 0 0 1 6 12zm10.8 3.59L16 14h-1l-1-2h-4l1-2h1l1-2h1l.68-1.36a6 6 0 0 1 2.12 9z': 'site',
    'M22 0H2a2 2 0 0 0-2 2v20a2 2 0 0 0 2 2h20a2 2 0 0 0 2-2V2a2 2 0 0 0-2-2zm-2.7 17h-2c-.64 0-.72-.47-1.9-1.6-1-1-1.42-1.06-1.66-1.06s-.37.17-.37.54v1.5c0 .45-.25.62-1.3.62A6.67 6.67 0 0 1 7 14c-2.14-2.93-2.69-5.1-2.69-5.57A.42.42 0 0 1 4.76 8h1.75c.46 0 .63.13.8.62.8 2.36 2.24 4.44 2.84 4.44.23 0 .25-.19.25-.67V10c0-1.17-.65-1.27-.65-1.69 0-.18.13-.3.35-.3h2.81c.38 0 .43.13.43.59v3.26c0 .37.08.51.24.51s.4-.12.77-.54a15.9 15.9 0 0 0 2.1-3.3.65.65 0 0 1 .67-.43h1.76c.35 0 .5.19.43.53a20.73 20.73 0 0 1-2.23 3.79c-.18.3-.26.46 0 .74s.78.77 1.14 1.25a4.77 4.77 0 0 1 1.47 2.17c0 .33-.19.42-.39.42z': 'vk',
    'M12 9.66A1.66 1.66 0 1 0 10.34 8 1.67 1.67 0 0 0 12 9.66z': 'ok',
    'M26.78 13.78a11.43 11.43 0 0 0-.64-2.06 10.55 10.55 0 0 0-1-1.87 11.61 11.61 0 0 0-1.34-1.63 11 11 0 0 0-1.63-1.34 10.56 10.56 0 0 0-1.87-1 10.81 10.81 0 0 0-2.07-.65 11.34 11.34 0 0 0-4.42 0 10.81 10.81 0 0 0-2.07.65 10.56 10.56 0 0 0-1.87 1 11 11 0 0 0-1.65 1.34 11.61 11.61 0 0 0-1.34 1.63L8.54 11A9.16 9.16 0 0 1 11 8.54a8.08 8.08 0 0 1 1.53-.83 9.1 9.1 0 0 1 1.68-.53 9.29 9.29 0 0 1 3.64 0 9.1 9.1 0 0 1 1.68.53 8.08 8.08 0 0 1 1.47.83A9.16 9.16 0 0 1 23.46 11a8.08 8.08 0 0 1 .83 1.53 9.1 9.1 0 0 1 .53 1.68A9.28 9.28 0 0 1 25 16a9.11 9.11 0 0 1-.18 1.81 9 9 0 0 1-.53 1.69 8.08 8.08 0 0 1-.83 1.5A9.16 9.16 0 0 1 21 23.46a8.08 8.08 0 0 1-1.53.83 9.1 9.1 0 0 1-1.68.53 9.29 9.29 0 0 1-3.64 0 9.1 9.1 0 0 1-1.68-.53 8.08 8.08 0 0 1-1.47-.83A9.16 9.16 0 0 1 8.54 21l-1.66 1.15a11.61 11.61 0 0 0 1.34 1.63 11 11 0 0 0 1.63 1.34 10.56 10.56 0 0 0 1.87 1 10.81 10.81 0 0 0 2.07.65 11.21 11.21 0 0 0 4.42 0 10.81 10.81 0 0 0 2.07-.65 10.56 10.56 0 0 0 1.87-1 11 11 0 0 0 1.63-1.34 11.61 11.61 0 0 0 1.34-1.63 10.55 10.55 0 0 0 1-1.87 11.43 11.43 0 0 0 .64-2.06 11.31 11.31 0 0 0 0-4.44z': 'message_xxx',
    'M15.793 9.4l1.414 1.414L12 16.024l-5.207-5.21L8.207 9.4 12 13.195z': 'description',
    'm10.758 6.03-.273-1.09a1.562 1.562 0 1 1 3.03 0l-.273 1.09a1.28 1.28 0 0 1-2.485 0ZM10 20v-5H9a1 1 0 0 1-1-1v-4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v4a1 1 0 0 1-1 1h-1v5a1 1 0 0 1-1 1h-2a1 1 0 0 1-1-1Z': 'visit_statistics',
}

data_info = {
            'city': '',
            'title': '',
            'email': '',
            'time_work': '',
            'address': '',
            'site': '',
            'ВКонтакте': '',
            'Одноклассники': '',
            'message_xxx': '',
            'description': '',
            'visit_statistics': '',
            'phone_1': '',
            'phone_2': '',
            'phone_3': '',
            'url': '',
        }
try:
    wb = load_workbook('info.xlsx')
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active
    ws.append(list(data_info.items()))

clean_text_lits = ['Показать вход', 'Открыто', 'Закрыто']

chrome_options = Options()
# chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_argument('--disable-notifications')
chrome_options.add_argument("--mute-audio")

# chrome_options.add_argument('--ignore-certificate-errors-spki-list') #handshake failed; returned -1, SSL error code 1, net_error -101


driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

print('Start Work Scripts')

base_url = "https://2gis.ru/{city}/search/Аптеки"

base_url = str(input(r"Вставьте ссылку заменив город на {city}"))

for city in city_list:
    base_url = base_url.replace(r"{city}", city)
    driver.get(base_url)
    # _1hf7139
    element_list = driver.find_elements(By.CLASS_NAME, "_1hf7139")
    count = 0
    skip = 0
    page = 1
    logger.info('Start script')
    for element in element_list:
        try:
            element.click()
        except:
            skip += 1
            logger.error(f'SKIP: {skip} / {len(element_list)}. Page: {page}')
            continue
        time.sleep(3)
        right_elem = driver.find_element(By.CLASS_NAME, "_18lzknl")
        title = right_elem.find_element(By.CLASS_NAME, "_oqoid").text
        count += 1
        logger.info(f'PARS: {count} / {len(element_list)}. Page: {page}. {title}')
        element_dict = data_info
        element_dict.update({
            'title': title,
            'city': city,
            'url': driver.current_url,
            })
        
        for el in right_elem.find_elements(By.CLASS_NAME, "_172gbf8"):
            svg_path = el.find_element(By.TAG_NAME, 'path').get_attribute('d')
            el_title = icon_dict.get(svg_path, svg_path)
            if el_title == 'phone':
                phone_el = el.find_elements(By.TAG_NAME, "a")
                for numb in range(len(phone_el)):
                    phone_href = str(phone_el[numb].get_attribute('href')).replace('tel:', '')
                    element_dict.update({f'phone_{numb+1}': phone_href})
            else:
                el_text = el.find_element(By.CLASS_NAME, "_49kxlr").text
                for clean in clean_text_lits:
                    el_text.replace(clean, '')
                element_dict.update({el_title: el_text})
        
        # Соц. Сети
        for el in right_elem.find_elements(By.CLASS_NAME, "_14uxmys"):
            el_href = el.find_element(By.TAG_NAME, "a").get_attribute('href')
            el_title = el.find_element(By.TAG_NAME, "a").get_attribute('aria-label')
            # Нормализация ссылки соц сети
            driver.switch_to.new_window('tab')
            driver.get(el_href)
            time.sleep(2)
            el_href = driver.current_url
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            element_dict.update({el_title: el_href})
        
        ws.append(list(element_dict.values()))
        wb.save('info.xlsx')
        time.sleep(random.randint(2, 6))
        
    time.sleep(3)


    driver.close()
driver.quit()