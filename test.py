from openpyxl import Workbook, load_workbook
import os

file_name = "example.xlsx"

if os.path.isfile(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Header 1", "Header 2"])  # добавьте заголовки, если создаёте новый файл

# Добавление данных в таблицу Excel
ws.append(["Data 1", "Data 2"])
ws.append(["Data 3", "Data 4"])

# Сохранение и закрытие файла Excel
wb.save(file_name)
# wb.close().